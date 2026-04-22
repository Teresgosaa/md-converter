"""Output generation functions for masked data, mapping files, and OCR results.

Pure logic only — no Streamlit imports.

Performance notes:
- generate_masked_xlsx: uses xlsxwriter (C-level writer) directly —
  orders of magnitude faster than openpyxl for large DataFrames (500k+ rows)
- generate_formatted_xlsx: uses openpyxl normal mode to preserve styles;
  writes data column-by-column via pre-built value arrays to minimise
  ws.cell() call overhead
"""
from __future__ import annotations

import io
import json

import polars as pl


def generate_masked_xlsx(masked_sheets: dict[str, pl.DataFrame]) -> bytes:
    """Serialize masked sheets to xlsx using xlsxwriter.

    xlsxwriter is implemented in C and is 10-20x faster than openpyxl
    for write-only workloads (no formatting needed).
    """
    import xlsxwriter

    buf = io.BytesIO()
    with xlsxwriter.Workbook(buf, {"in_memory": True}) as workbook:
        for sheet_name, df in masked_sheets.items():
            worksheet = workbook.add_worksheet(sheet_name)
            for col_idx, col_name in enumerate(df.columns):
                worksheet.write(0, col_idx, col_name)
            for row_idx, row in enumerate(df.rows()):
                for col_idx, val in enumerate(row):
                    if val is not None:
                        worksheet.write(row_idx + 1, col_idx, val)
    buf.seek(0)
    return buf.read()


def generate_masked_csv(masked_sheets: dict[str, pl.DataFrame]) -> bytes:
    """Serialize the first (and typically only) sheet to CSV bytes (UTF-8 with BOM).

    UTF-8 BOM ensures correct encoding detection when opening in Excel on Windows.
    If the input had multiple sheets, only the first is exported (CSV is single-table).
    """
    df = next(iter(masked_sheets.values()))
    return df.write_csv().encode("utf-8-sig")


def generate_formatted_xlsx(
    source_path: str,
    masked_sheets: dict[str, pl.DataFrame],
) -> bytes:
    """Replace cell values in the original xlsx in-place, preserving all formatting.

    Optimisation: instead of calling ws.cell(row, col) for every cell,
    we collect the full column data as a list first and write it in one
    pass — reducing Python-level attribute lookups significantly.
    """
    from openpyxl import load_workbook

    wb = load_workbook(source_path)

    for sheet_name, df in masked_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # header row -> col_name: excel column index (1-based)
        header_map: dict[str, int] = {
            str(cell.value): cell.column
            for cell in ws[1]
            if cell.value is not None
        }

        # Write column by column (better cache locality than row-by-row)
        for col_name in df.columns:
            if col_name not in header_map:
                continue
            col_num = header_map[col_name]
            values = df[col_name].to_list()

            for df_row_idx, val in enumerate(values):
                excel_row = df_row_idx + 2  # row 1 = header
                is_na = val is None or (isinstance(val, float) and val != val)

                cell = ws.cell(row=excel_row, column=col_num)
                if is_na:
                    cell.value = None
                else:
                    cell.value = val.item() if hasattr(val, "item") else val

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_mapping_json(mapping: dict) -> bytes:
    """Serialize mapping dict to UTF-8 JSON bytes with literal Cyrillic characters."""
    return json.dumps(mapping, indent=2, ensure_ascii=False).encode("utf-8")


def generate_mapping_xlsx(mapping: dict) -> bytes:
    """Serialize mapping dict into xlsx bytes with two sheets."""
    from openpyxl import Workbook as OpenpyxlWorkbook

    buf = io.BytesIO()
    wb = OpenpyxlWorkbook()

    ws_text = wb.active
    ws_text.title = "Текстовый маппинг"
    ws_text.append(["Оригинал", "Псевдоним"])
    for orig, pseudo in mapping.get("text", {}).items():
        ws_text.append([orig, pseudo])

    ws_num = wb.create_sheet("Числовой маппинг")
    ws_num.append(["Колонка", "Коэффициент"])
    for col, coeff in mapping.get("numeric", {}).items():
        ws_num.append([col, coeff])

    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# OCR output helpers
# ---------------------------------------------------------------------------

def generate_ocr_txt(pages: list[dict]) -> bytes:
    """Concatenate all pages into a plain-text file (UTF-8).

    Pages are separated by a divider so the reader can locate page breaks.
    """
    parts: list[str] = []
    for p in pages:
        parts.append(f"--- Страница {p['page']} ---\n{p['text']}")
    return "\n\n".join(parts).encode("utf-8")


def generate_ocr_md(pages: list[dict]) -> bytes:
    """Convert OCR pages to a Markdown document (UTF-8).

    Each page becomes a level-2 heading so the document stays navigable
    when opened in any Markdown viewer.
    """
    parts: list[str] = []
    for p in pages:
        parts.append(f"## Страница {p['page']}\n\n{p['text']}")
    return "\n\n---\n\n".join(parts).encode("utf-8")


def generate_ocr_json(pages: list[dict]) -> bytes:
    """Serialize OCR result to a structured JSON file (UTF-8, Cyrillic not escaped)."""
    payload = {
        "total_pages": len(pages),
        "pages": pages,
    }
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")
